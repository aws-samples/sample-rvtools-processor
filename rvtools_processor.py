import openpyxl
import pandas as pd
import sys
import os
import argparse
from pathlib import Path
import json
from datetime import datetime
import warnings

# Suppress pandas warnings
warnings.simplefilter(action='ignore', category=FutureWarning)

def find_rvtools_files(directory="."):
 """Find RVTools files in directory"""
 files = []
 xlsx_files = list(Path(directory).glob("*.xlsx"))
 
 for file in xlsx_files:
 if file.name.startswith("~"):
 continue
 try:
 wb = openpyxl.load_workbook(file, read_only=True)
 rvtools_sheets = {'vInfo', 'vHost', 'vCluster'}
 if any(sheet in wb.sheetnames for sheet in rvtools_sheets):
 files.append(file)
 except Exception:
 continue
 
 return files

def generate_output_filename(mode, input_files=None):
 """Generate output filename with timestamp"""
 timestamp = datetime.now().strftime("%Y%m%d_%H%M")
 
 if mode == "both":
 return f"RVTools_Consolidated_Anonymized_{timestamp}.xlsx"
 elif mode == "consolidate":
 return f"RVTools_Combined_{timestamp}.xlsx"
 elif mode == "anonymize":
 return f"RVTools_Anonymized_{timestamp}.xlsx"
 elif mode == "deanonymize":
 return f"RVTools_Deanonymized_{timestamp}.xlsx"
def consolidate_rvtools(input_files, output_file):
 """Consolidate multiple RVTools files into one"""
 print(f"\nStarting consolidation of {len(input_files)} files...")
 consolidated_sheets = {}
 
 for file_path in input_files:
 print(f"Processing: {file_path.name}")
 try:
 wb = openpyxl.load_workbook(filename=file_path, read_only=True)
 for sheet_name in wb.sheetnames:
 df = pd.read_excel(file_path, sheet_name=sheet_name)
 if sheet_name not in consolidated_sheets:
 consolidated_sheets[sheet_name] = df
 else:
 consolidated_sheets[sheet_name] = pd.concat([consolidated_sheets[sheet_name], df], 
 ignore_index=True)
 wb.close()
 except Exception as e:
 print(f"Error processing {file_path.name}: {str(e)}")
 continue

 print(f"Creating consolidated file: {output_file}")
 with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
 for sheet_name, df in consolidated_sheets.items():
 df.to_excel(writer, sheet_name=sheet_name, index=False)

 print("\nConsolidation Summary:")
 print(f"Input files processed: {len(input_files)}")
 print("Sheets consolidated:")
 for sheet_name, df in consolidated_sheets.items():
 print(f" - {sheet_name}: {len(df)} total rows")
 
 return output_file

def create_mapping(original_value, vm_id, mapping_dict):
 """Create mapping between original and anonymized values"""
 if not original_value or not vm_id:
 return None
 
 anonymized = str(vm_id)
 mapping_dict[anonymized] = str(original_value)
 return anonymized
def anonymize_rvtools(input_file, output_file):
 """Anonymize RVTools file using VM ID"""
 print(f"\nStarting anonymization...")
 
 mapping_dict = {}
 sensitive_fields = {
 # vInfo basic fields
 'Primary IP Address': '***************',
 'min Required EVC Mode': 'ANON_EVC',
 'Folder': 'ANON_FOLDER',
 'Datacenter': 'ANON_DC',
 'Cluster': 'ANON_CLUSTER',
 'VI SDK Server': 'ANON_SDK',
 
 # Disk related
 'Disk Path': 'ANON_DISK_PATH',
 'Internal Sort Column': 'ANON_SORT',
 'Mac Address': 'ANON_MAC',
 
 # VM specific fields
 'VM': 'ID',
 'DNS Name': 'ANON_DNS',
 'Resource Pool': 'ANON_POOL',
 'Path': 'ANON_PATH',
 'Log directory': 'ANON_LOG',
 'Snapshot directory': 'ANON_SNAP',
 'Suspend directory': 'ANON_SUSP',
 'Annotation': 'ANON_NOTE',
 'Host': 'ANON_HOST',
 'IP Address': '***************'
 }
 
 try:
 wb = openpyxl.load_workbook(filename=input_file)
 
 for sheet_name in wb.sheetnames:
 sheet = wb[sheet_name]
 headers = [cell.value for cell in sheet[1]]
 
 for row in sheet.iter_rows(min_row=2):
 for idx, cell in enumerate(row):
 header = headers[idx]
 if header in sensitive_fields:
 if header in ['IP Address', 'Primary IP Address', 'Mac Address']:
 cell.value = sensitive_fields[header]
 else:
 vm_id_col = headers.index('VM ID') if 'VM ID' in headers else None
 if vm_id_col is not None:
 vm_id = row[vm_id_col].value
 original_value = cell.value
 cell.value = create_mapping(original_value, vm_id, mapping_dict)
 else:
 cell.value = f"{sensitive_fields[header]}_UNKNOWN"
 
 # Save mapping file
 mapping_file = f"mapping_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
 with open(mapping_file, 'w') as f:
 json.dump(mapping_dict, f, indent=2)
 
 wb.save(output_file)
 print(f"Created mapping file: {mapping_file}")
 print(f"Created anonymized file: {output_file}")
 return output_file, mapping_file
 
 except Exception as e:
 print(f"Error during anonymization: {str(e)}")
 return None, None

def deanonymize_rvtools(input_file, mapping_file, output_file):
 """Restore original names using mapping file"""
 print(f"\nStarting deanonymization...")
 
 try:
 # Load mapping
 with open(mapping_file, 'r') as f:
 mapping = json.load(f)
 
 wb = openpyxl.load_workbook(filename=input_file)
 
 for sheet_name in wb.sheetnames:
 sheet = wb[sheet_name]
 headers = [cell.value for cell in sheet[1]]
 
 for row in sheet.iter_rows(min_row=2):
 for cell in row:
 if cell.value and str(cell.value) in mapping:
 cell.value = mapping[str(cell.value)]
 
 wb.save(output_file)
 print(f"Created deanonymized file: {output_file}")
 return output_file
 
 except Exception as e:
 print(f"Error during deanonymization: {str(e)}")
 return None
def process_rvtools(args):
 """Main processing function supporting all modes"""
 input_files = []
 
 if not hasattr(args, 'input_files') or not args.input_files:
 input_files = find_rvtools_files()
 if not input_files:
 raise ValueError("No RVTools files found in current directory")
 print(f"Found {len(input_files)} RVTools files: {[f.name for f in input_files]}")
 else:
 input_files = [Path(f) for f in args.input_files]

 if not hasattr(args, 'output') or not args.output:
 args.output = generate_output_filename(args.mode, input_files)

 if args.mode == "consolidate":
 return consolidate_rvtools(input_files, args.output)
 elif args.mode == "anonymize":
 return anonymize_rvtools(input_files[0], args.output)
 elif args.mode == "deanonymize":
 if not args.mapping:
 raise ValueError("Mapping file required for deanonymization")
 return deanonymize_rvtools(input_files[0], args.mapping, args.output)
 elif args.mode == "both":
 temp_file = f"temp_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
 consolidated = consolidate_rvtools(input_files, temp_file)
 result, mapping = anonymize_rvtools(consolidated, args.output)
 os.remove(temp_file)
 return result

def main():
 parser = argparse.ArgumentParser(description="RVTools Processing Tool")
 subparsers = parser.add_subparsers(dest='mode')
 
 # Consolidate parser
 parser_consolidate = subparsers.add_parser('consolidate',
 help='Consolidate multiple RVTools files')
 parser_consolidate.add_argument('input_files', nargs='*',
 help='Input files (optional, will find all RVTools files if not specified)')
 parser_consolidate.add_argument('-o', '--output',
 help='Output filename (optional)')

 # Anonymize parser
 parser_anonymize = subparsers.add_parser('anonymize',
 help='Anonymize RVTools file')
 parser_anonymize.add_argument('input_files', nargs='*',
 help='Input file (optional, will find RVTools files if not specified)')
 parser_anonymize.add_argument('-o', '--output',
 help='Output filename (optional)')

 # Deanonymize parser
 parser_deanonymize = subparsers.add_parser('deanonymize',
 help='Deanonymize RVTools file')
 parser_deanonymize.add_argument('input_files', nargs='*',
 help='Input anonymized file')
 parser_deanonymize.add_argument('-m', '--mapping', required=True,
 help='Mapping file from anonymization')
 parser_deanonymize.add_argument('-o', '--output',
 help='Output filename (optional)')

 # Both consolidate and anonymize
 parser_both = subparsers.add_parser('both',
 help='Consolidate and anonymize RVTools files')
 parser_both.add_argument('input_files', nargs='*',
 help='Input files (optional, will find all RVTools files if not specified)')
 parser_both.add_argument('-o', '--output',
 help='Output filename (optional)')

 args = parser.parse_args()
 
 if not args.mode:
 parser.print_help()
 sys.exit(1)
 
 try:
 process_rvtools(args)
 except Exception as e:
 print(f"Error: {str(e)}")
 sys.exit(1)

if __name__ == "__main__":
 main()
